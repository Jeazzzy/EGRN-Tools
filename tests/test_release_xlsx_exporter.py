from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from core.release_checker import PdfAreaResult, STATUS_FOUND
from core.release_xlsx_exporter import export_release_results
from core.release_xml_checker import STATUS_VALID, XmlReleaseResult


def pdf_result():
    return PdfAreaResult(
        settlement="п. Виноградный",
        file_name="description.pdf",
        relative_path=r"п. Виноградный\description.pdf",
        full_path=r"C:\release\pdf\п. Виноградный\description.pdf",
        area="527210",
        status=STATUS_FOUND,
        details="",
    )


def xml_result(object_name="Территориальная зона ВИ1"):
    return XmlReleaseResult(
        settlement_folder="Вне НП",
        zone_folder="ВИ1",
        archive_name="boundaries.zip",
        xml_name="boundaries.xml",
        object_name=object_name,
        cadastral_district="23",
        index="ВИ1",
        locality="Вне НП",
        point_count=7,
        polygon_count=2,
        checked_accuracy_points=9,
        accuracy_error_count=0,
        accuracy_summary="ОК (9)",
        status=STATUS_VALID,
        details="",
        full_path=r"C:\release\xml\Вне НП\ВИ1\boundaries.zip",
    )


class ReleaseXlsxExporterTests(unittest.TestCase):
    def test_exports_pdf_and_xml_to_separate_sheets(self):
        with TemporaryDirectory() as temporary:
            path = export_release_results(
                Path(temporary) / "release-check",
                [pdf_result()],
                [xml_result()],
            )

            self.assertEqual(path.suffix, ".xlsx")
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames, ["PDF", "XML"])

            pdf_sheet = workbook["PDF"]
            self.assertEqual(pdf_sheet["A1"].value, "Населённый пункт")
            self.assertEqual(pdf_sheet["C2"].value, 527210)
            self.assertEqual(pdf_sheet.freeze_panes, "A2")
            self.assertTrue(pdf_sheet.auto_filter.ref)

            xml_sheet = workbook["XML"]
            self.assertEqual(xml_sheet["E2"].value, "Территориальная зона ВИ1")
            self.assertEqual(xml_sheet["I2"].value, 7)
            self.assertEqual(xml_sheet["J2"].value, 2)
            self.assertEqual(xml_sheet["N2"].value, STATUS_VALID)

    def test_exports_only_available_result_type(self):
        with TemporaryDirectory() as temporary:
            path = export_release_results(
                Path(temporary) / "xml-only.xlsx",
                xml_results=[xml_result()],
            )

            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames, ["XML"])

    def test_neutralises_formula_text_from_xml(self):
        with TemporaryDirectory() as temporary:
            path = export_release_results(
                Path(temporary) / "safe.xlsx",
                xml_results=[xml_result("=1+1")],
            )

            workbook = load_workbook(path, data_only=False)
            self.assertEqual(workbook["XML"]["E2"].value, "'=1+1")
            self.assertEqual(workbook["XML"]["E2"].data_type, "s")

    def test_rejects_export_without_results(self):
        with TemporaryDirectory() as temporary:
            with self.assertRaisesRegex(ValueError, "Нет результатов"):
                export_release_results(Path(temporary) / "empty.xlsx")


if __name__ == "__main__":
    unittest.main()
