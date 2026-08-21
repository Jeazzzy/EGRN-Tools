"""Экспорт результатов проверки выпуска в книгу XLSX."""

from __future__ import annotations

from pathlib import Path
import re
from typing import Iterable

from .release_checker import (
    STATUS_FOUND,
    STATUS_MULTIPLE,
    PdfAreaResult,
)
from .release_xml_checker import (
    STATUS_INVALID,
    STATUS_READ_ERROR,
    STATUS_VALID,
    XmlReleaseResult,
)


PDF_HEADERS = (
    "Населённый пункт",
    "PDF-файл",
    "Площадь, м²",
    "Страниц",
    "Статус",
    "Примечание",
    "Полный путь",
)

XML_HEADERS = (
    "Папка НП",
    "Папка зоны",
    "ZIP",
    "XML",
    "Полное название объекта",
    "Кадастровый район",
    "Индекс",
    "НП в XML",
    "Тип границы",
    "Реестровый номер",
    "Точек",
    "Контуров всего",
    "Внешних контуров",
    "Внутренних (дырок)",
    "Проверено точек",
    "Ошибок точности",
    "Проверка точности",
    "Статус",
    "Ошибки",
    "Полный путь",
)


def _safe_excel_value(value):
    """Не позволяет тексту из входных файлов превратиться в формулу Excel."""

    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _pdf_rows(results: Iterable[PdfAreaResult]):
    for result in results:
        area = (
            int(result.area)
            if re.fullmatch(r"[0-9]+", result.area or "")
            else result.area
        )
        yield (
            result.settlement,
            result.file_name,
            area,
            result.page_count,
            result.status,
            result.details,
            result.full_path,
        )


def _xml_rows(results: Iterable[XmlReleaseResult]):
    for result in results:
        yield (
            result.settlement_folder,
            result.zone_folder,
            result.archive_name,
            result.xml_name,
            result.object_name,
            result.cadastral_district,
            result.index,
            result.locality,
            result.boundary_type,
            result.registry_number,
            result.point_count,
            result.polygon_count,
            result.outer_contour_count,
            result.hole_count,
            result.checked_accuracy_points,
            result.accuracy_error_count,
            result.accuracy_summary,
            result.status,
            result.details,
            result.full_path,
        )


def _add_sheet(workbook, title, headers, rows, status_column, status_colours):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(tuple(_safe_excel_value(value) for value in row))

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    fills = {
        status: PatternFill("solid", fgColor=colour)
        for status, colour in status_colours.items()
    }
    fallback_fill = PatternFill("solid", fgColor="FEE2E2")
    for row_number in range(2, sheet.max_row + 1):
        status = sheet.cell(row=row_number, column=status_column).value
        row_fill = fills.get(status, fallback_fill)
        for cell in sheet[row_number]:
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        values = [
            str(sheet.cell(row=row, column=column).value or "")
            for row in range(1, sheet.max_row + 1)
        ]
        maximum = max(
            (len(line) for value in values for line in value.splitlines()),
            default=0,
        )
        sheet.column_dimensions[get_column_letter(column)].width = min(
            max(12, maximum + 2),
            60,
        )
    return sheet


def export_release_results(
    output_path: str | Path,
    pdf_results: Iterable[PdfAreaResult] = (),
    xml_results: Iterable[XmlReleaseResult] = (),
) -> Path:
    """Создаёт XLSX с отдельными листами для имеющихся результатов."""

    try:
        from openpyxl import Workbook
    except ImportError as error:
        raise RuntimeError(
            "Не установлен модуль openpyxl. Установите зависимости "
            "из requirements.txt."
        ) from error

    pdf_results = list(pdf_results)
    xml_results = list(xml_results)
    if not pdf_results and not xml_results:
        raise ValueError("Нет результатов для экспорта.")

    path = Path(output_path)
    if path.suffix.casefold() != ".xlsx":
        path = path.with_suffix(".xlsx")
    if not path.parent.is_dir():
        raise ValueError("Папка для сохранения XLSX не существует.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "Проверка выпуска"
    workbook.properties.subject = "Результаты проверки PDF и XML"

    if pdf_results:
        _add_sheet(
            workbook,
            "PDF",
            PDF_HEADERS,
            _pdf_rows(pdf_results),
            status_column=5,
            status_colours={
                STATUS_FOUND: "DCFCE7",
                STATUS_MULTIPLE: "FEF3C7",
            },
        )
    if xml_results:
        _add_sheet(
            workbook,
            "XML",
            XML_HEADERS,
            _xml_rows(xml_results),
            status_column=18,
            status_colours={
                STATUS_VALID: "DCFCE7",
                STATUS_INVALID: "FEF3C7",
                STATUS_READ_ERROR: "FEE2E2",
            },
        )

    workbook.active = 0
    workbook.save(path)
    return path
